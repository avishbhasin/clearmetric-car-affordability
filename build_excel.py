"""
ClearMetric Car Affordability Calculator — Premium Excel Template
Product T8 for Gumroad ($9.99)

3 Sheets:
  1. Affordability Calculator — inputs + max price + monthly costs + 20/4/10 rule
  2. Total Cost Comparison — compare 3 cars side by side
  3. How To Use — instructions

Design: Charcoal/Silver palette (#2C3E50 primary, #1C2833 dark, #D5D8DC input)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import os

# ============================================================
# DESIGN SYSTEM — Charcoal/Silver
# ============================================================
PRIMARY = "2C3E50"
DARK = "1C2833"
WHITE = "FFFFFF"
INPUT_TINT = "D5D8DC"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="AEB6BF", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=PRIMARY, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=PRIMARY)
FONT_SMALL = Font(name="Calibri", size=9, color=MED_GRAY, italic=True)

FILL_PRIMARY = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY),
    right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY),
    bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_PRIMARY
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_PRIMARY
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: AFFORDABILITY CALCULATOR
# ============================================================
def build_affordability(ws):
    ws.title = "Affordability Calculator"
    ws.sheet_properties.tabColor = PRIMARY
    cols(ws, {"A": 2, "B": 32, "C": 18, "D": 4, "E": 32, "F": 18})

    for r in range(1, 55):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="CAR AFFORDABILITY CALCULATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:E3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Enter your numbers in the gray cells. Results update automatically.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== LEFT: INPUTS =====
    header_bar(ws, 5, 2, 3, "INCOME & DEBT")
    label_input(ws, 6, 2, 3, "Monthly Take-Home Income", 5000, "$#,##0")
    label_input(ws, 7, 2, 3, "Monthly Debt Payments", 500, "$#,##0")
    label_input(ws, 8, 2, 3, "Max % Income for Car (10%=20/4/10)", 0.15, "0%")

    header_bar(ws, 10, 2, 3, "DOWN PAYMENT & TRADE-IN")
    label_input(ws, 11, 2, 3, "Down Payment", 5000, "$#,##0")
    label_input(ws, 12, 2, 3, "Trade-In Value", 0, "$#,##0")

    header_bar(ws, 14, 2, 3, "LOAN TERMS")
    label_input(ws, 15, 2, 3, "Loan Term (months)", 60, "0")
    label_input(ws, 16, 2, 3, "Interest Rate", 0.065, "0.00%")
    label_input(ws, 17, 2, 3, "Sales Tax Rate", 0.07, "0.0%")

    header_bar(ws, 19, 2, 3, "OTHER COSTS")
    label_input(ws, 20, 2, 3, "Monthly Insurance", 150, "$#,##0")
    label_input(ws, 21, 2, 3, "Annual Maintenance", 1200, "$#,##0")
    label_input(ws, 22, 2, 3, "Annual Fuel Cost", 2400, "$#,##0")

    # ===== RIGHT: RESULTS =====
    header_bar(ws, 5, 5, 6, "RESULTS", FILL_DARK)

    # Max affordable monthly payment
    label_calc(ws, 6, 5, 6, "Max Monthly Payment (P&I)",
               "=MAX(0,C6*C8-C7)", "$#,##0", bold=True)
    # Max loan amount (PV of max payment)
    label_calc(ws, 7, 5, 6, "Max Loan Amount",
               "=IF(C16>0,F6*((1-(1+C16/12)^(-C15))/(C16/12)),F6*C15)", "$#,##0")
    # Max car price
    label_calc(ws, 8, 5, 6, "Max Car Price (before tax)",
               "=(F7+C11+C12)/(1+C17)", "$#,##0", bold=True)

    # Monthly costs
    header_bar(ws, 10, 5, 6, "MONTHLY COSTS")
    label_calc(ws, 11, 5, 6, "Loan Payment", "=F6", "$#,##0")
    label_calc(ws, 12, 5, 6, "Insurance", "=C20", "$#,##0")
    label_calc(ws, 13, 5, 6, "Maintenance (mo)", "=C21/12", "$#,##0")
    label_calc(ws, 14, 5, 6, "Fuel (mo)", "=C22/12", "$#,##0")
    label_calc(ws, 15, 5, 6, "True Monthly Cost",
               "=F6+C20+C21/12+C22/12", "$#,##0", bold=True)

    # 20/4/10 rule
    header_bar(ws, 17, 5, 6, "20/4/10 RULE CHECK")
    ws.merge_cells("E18:F18")
    ws.cell(row=18, column=5, value='=IF(AND(C11>=F8*0.2,C15<=48,F15<=C6*0.1),"PASS","FAIL")')
    ws.cell(row=18, column=5).font = FONT_BOLD
    ws.cell(row=18, column=5).fill = FILL_WHITE
    ws.cell(row=18, column=5).border = THIN
    ws.cell(row=18, column=5).alignment = ALIGN_C
    for c in range(5, 7):
        ws.cell(row=18, column=c).border = THIN

    # Unlock input cells
    ws.protection.sheet = True
    input_cells = [(6, 3), (7, 3), (8, 3), (11, 3), (12, 3), (15, 3), (16, 3), (17, 3),
                   (20, 3), (21, 3), (22, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: TOTAL COST COMPARISON
# ============================================================
def build_comparison(wb):
    ws = wb.create_sheet("Total Cost Comparison")
    ws.sheet_properties.tabColor = "5D6D7E"
    fc = "'Affordability Calculator'"
    cols(ws, {"A": 2, "B": 24, "C": 18, "D": 18, "E": 4})

    for r in range(1, 45):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 5):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:E1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="TOTAL COST COMPARISON — 3 CARS").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:E3")
    ws.cell(row=3, column=2, value="Compare 3 cars. Enter price, loan, insurance, fuel, maintenance in gray cells.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Headers
    headers = ["", "Car 1", "Car 2", "Car 3"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_PRIMARY
        cell.alignment = ALIGN_C
        cell.border = THIN

    # Row labels
    labels = ["Car Price", "Loan Amount", "Monthly Payment", "Annual Insurance",
              "Annual Fuel", "Annual Maintenance", "5-Yr Depreciation (est)", "5-Yr Total Cost"]
    for i, lbl in enumerate(labels):
        r = 6 + i
        ws.cell(row=r, column=2, value=lbl).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).alignment = ALIGN_L

    # Car 1 (col C)
    ws.cell(row=6, column=3, value=25000).fill = FILL_INPUT
    ws.cell(row=6, column=3).font = FONT_INPUT
    ws.cell(row=6, column=3).number_format = "$#,##0"
    ws.cell(row=7, column=3, value="=C6*0.8").fill = FILL_INPUT
    ws.cell(row=7, column=3).font = FONT_INPUT
    ws.cell(row=7, column=3).number_format = "$#,##0"
    ws.cell(row=8, column=3, value=f"=IF({fc}!C16>0,PMT({fc}!C16/12,{fc}!C15,-C7),C7/{fc}!C15)")
    ws.cell(row=8, column=3).font = FONT_BOLD
    ws.cell(row=8, column=3).number_format = "$#,##0"
    ws.cell(row=9, column=3, value=1800).fill = FILL_INPUT
    ws.cell(row=9, column=3).font = FONT_INPUT
    ws.cell(row=9, column=3).number_format = "$#,##0"
    ws.cell(row=10, column=3, value=2400).fill = FILL_INPUT
    ws.cell(row=10, column=3).font = FONT_INPUT
    ws.cell(row=10, column=3).number_format = "$#,##0"
    ws.cell(row=11, column=3, value=1200).fill = FILL_INPUT
    ws.cell(row=11, column=3).font = FONT_INPUT
    ws.cell(row=11, column=3).number_format = "$#,##0"
    ws.cell(row=12, column=3, value="=C6*(1-0.8*0.85^4)")
    ws.cell(row=12, column=3).font = FONT_BOLD
    ws.cell(row=12, column=3).number_format = "$#,##0"
    ws.cell(row=13, column=3, value="=C8*60+C9*5+C10*5+C11*5+C12")
    ws.cell(row=13, column=3).font = FONT_BOLD
    ws.cell(row=13, column=3).number_format = "$#,##0"
    for r in range(6, 14):
        ws.cell(row=r, column=3).border = THIN
        ws.cell(row=r, column=3).alignment = ALIGN_R

    # Car 2 (col D)
    ws.cell(row=6, column=4, value=30000).fill = FILL_INPUT
    ws.cell(row=6, column=4).font = FONT_INPUT
    ws.cell(row=6, column=4).number_format = "$#,##0"
    ws.cell(row=7, column=4, value="=D6*0.8").fill = FILL_INPUT
    ws.cell(row=7, column=4).font = FONT_INPUT
    ws.cell(row=7, column=4).number_format = "$#,##0"
    ws.cell(row=8, column=4, value=f"=IF({fc}!C16>0,PMT({fc}!C16/12,{fc}!C15,-D7),D7/{fc}!C15)")
    ws.cell(row=8, column=4).font = FONT_BOLD
    ws.cell(row=8, column=4).number_format = "$#,##0"
    ws.cell(row=9, column=4, value=2000).fill = FILL_INPUT
    ws.cell(row=9, column=4).font = FONT_INPUT
    ws.cell(row=9, column=4).number_format = "$#,##0"
    ws.cell(row=10, column=4, value=2400).fill = FILL_INPUT
    ws.cell(row=10, column=4).font = FONT_INPUT
    ws.cell(row=10, column=4).number_format = "$#,##0"
    ws.cell(row=11, column=4, value=1500).fill = FILL_INPUT
    ws.cell(row=11, column=4).font = FONT_INPUT
    ws.cell(row=11, column=4).number_format = "$#,##0"
    ws.cell(row=12, column=4, value="=D6*(1-0.8*0.85^4)")
    ws.cell(row=12, column=4).font = FONT_BOLD
    ws.cell(row=12, column=4).number_format = "$#,##0"
    ws.cell(row=13, column=4, value="=D8*60+D9*5+D10*5+D11*5+D12")
    ws.cell(row=13, column=4).font = FONT_BOLD
    ws.cell(row=13, column=4).number_format = "$#,##0"
    for r in range(6, 14):
        ws.cell(row=r, column=4).border = THIN
        ws.cell(row=r, column=4).alignment = ALIGN_R

    # Car 3 (col E)
    ws.cell(row=6, column=5, value=35000).fill = FILL_INPUT
    ws.cell(row=6, column=5).font = FONT_INPUT
    ws.cell(row=6, column=5).number_format = "$#,##0"
    ws.cell(row=7, column=5, value="=E6*0.8").fill = FILL_INPUT
    ws.cell(row=7, column=5).font = FONT_INPUT
    ws.cell(row=7, column=5).number_format = "$#,##0"
    ws.cell(row=8, column=5, value=f"=IF({fc}!C16>0,PMT({fc}!C16/12,{fc}!C15,-E7),E7/{fc}!C15)")
    ws.cell(row=8, column=5).font = FONT_BOLD
    ws.cell(row=8, column=5).number_format = "$#,##0"
    ws.cell(row=9, column=5, value=2200).fill = FILL_INPUT
    ws.cell(row=9, column=5).font = FONT_INPUT
    ws.cell(row=9, column=5).number_format = "$#,##0"
    ws.cell(row=10, column=5, value=2400).fill = FILL_INPUT
    ws.cell(row=10, column=5).font = FONT_INPUT
    ws.cell(row=10, column=5).number_format = "$#,##0"
    ws.cell(row=11, column=5, value=1800).fill = FILL_INPUT
    ws.cell(row=11, column=5).font = FONT_INPUT
    ws.cell(row=11, column=5).number_format = "$#,##0"
    ws.cell(row=12, column=5, value="=E6*(1-0.8*0.85^4)")
    ws.cell(row=12, column=5).font = FONT_BOLD
    ws.cell(row=12, column=5).number_format = "$#,##0"
    ws.cell(row=13, column=5, value="=E8*60+E9*5+E10*5+E11*5+E12")
    ws.cell(row=13, column=5).font = FONT_BOLD
    ws.cell(row=13, column=5).number_format = "$#,##0"
    for r in range(6, 14):
        ws.cell(row=r, column=5).border = THIN
        ws.cell(row=r, column=5).alignment = ALIGN_R

    ws.protection.sheet = True
    for r in [6, 7, 9, 10, 11]:
        for c in [3, 4, 5]:
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 3: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = MED_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE CAR AFFORDABILITY CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Affordability Calculator' tab and enter your numbers in the GRAY cells",
            "2. Results appear on the right: max car price, monthly payment, 20/4/10 rule",
            "3. Check the 'Total Cost Comparison' tab to compare 3 cars side by side",
        ]),
        ("INPUT EXPLANATIONS", [
            "Monthly Take-Home Income: Your net pay after taxes and deductions",
            "Monthly Debt Payments: Credit cards, student loans, other car payments, etc.",
            "Max % Income for Car: 10% = 20/4/10 rule; 15% is common; 20%+ is aggressive",
            "Down Payment: Cash you have for a down payment",
            "Trade-In Value: Estimated value of your current vehicle",
            "Loan Term: 36–72 months. Shorter = less interest, higher payment",
            "Interest Rate: Depends on credit score. Excellent 720+ ~5.5%, Good 670–719 ~6.5%, Fair 580–669 ~9%, Poor <580 ~12%",
            "Sales Tax Rate: Your state/local rate (e.g., 7% = 0.07)",
            "Monthly Insurance: Get a quote for the car you're considering",
            "Annual Maintenance: Oil changes, tires, repairs. Rule of thumb: $1,200–$2,000/year",
            "Annual Fuel Cost: Miles × (fuel price / MPG) × 12",
        ]),
        ("20/4/10 RULE", [
            "20% down: Put at least 20% down to avoid being underwater",
            "4-year loan: Finance for 4 years or less",
            "10% of income: Total car costs (payment + insurance + maintenance + fuel) ≤ 10% of income",
            "Passing all three = financially responsible car purchase",
        ]),
        ("TOTAL COST COMPARISON", [
            "Enter price, loan, insurance, fuel, maintenance for up to 3 cars",
            "Depreciation uses 20% first year, 15%/year after (typical for new cars)",
            "5-year total cost includes payments, insurance, fuel, maintenance, depreciation",
        ]),
        ("IMPORTANT NOTES", [
            "This calculator is for educational purposes only. Not financial advice.",
            "Rates and terms vary by lender. Get pre-approved before shopping.",
            "© 2026 ClearMetric. For personal use only.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=PRIMARY)
        ws.cell(row=r, column=2).fill = PatternFill(start_color=INPUT_TINT, end_color=INPUT_TINT, fill_type="solid")
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Affordability Calculator sheet...")
    build_affordability(ws)

    print("Building Total Cost Comparison sheet...")
    build_comparison(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "output",
        "ClearMetric-Car-Affordability-Calculator.xlsx",
    )
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
